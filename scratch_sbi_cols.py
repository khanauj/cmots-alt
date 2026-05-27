import openpyxl

wb = openpyxl.load_workbook("scratch_sbi_infra_portfolio.xlsx", read_only=True)
ws = wb.active

row05 = [cell.value for cell in ws[6]]  # row index 6 is line 5 (1-based index)
row10 = [cell.value for cell in ws[11]] # row index 11 is line 10 (1-based index)

# Print up to the first 15 columns
print("Row 05 (header):", row05[:15])
print("Row 10 (data):  ", row10[:15])
