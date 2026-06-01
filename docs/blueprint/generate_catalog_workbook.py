"""Phase 7 — generate the CMOTS-replacement API catalog workbook.

Single source of truth for Phase 6 (API design) AND Phase 7 (Excel output).
Produces a 3-sheet .xlsx:

    Sheet 1: EQUITY
    Sheet 2: MUTUAL FUNDS
    Sheet 3: CMOTS COMMENTS

Run:
    python docs/blueprint/generate_catalog_workbook.py
    -> docs/blueprint/CMOTS_Replacement_API_Catalog.xlsx

The catalog rows below are deliberately exhaustive. Columns per API row:
    api_number, report_name, api_url, frequency, updation_time,
    input, input_desc, output, data_type, output_desc
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Base URL used in every catalogued endpoint (matches the existing FastAPI app;
# routers are mounted at root today, versioned here for the public contract).
BASE = "https://api.cmots-alt.local/api/v1"

# ---------------------------------------------------------------------------
# API column order (Sheets 1 & 2)
# ---------------------------------------------------------------------------
API_COLUMNS = [
    ("api_number", "API Number", 12),
    ("report_name", "Report Name", 30),
    ("api_url", "API URL", 46),
    ("frequency", "Frequency", 16),
    ("updation_time", "Updation Time", 18),
    ("input", "Input", 22),
    ("input_desc", "Input Description", 34),
    ("output", "Output", 50),
    ("data_type", "Data Type", 14),
    ("output_desc", "Output Description", 50),
]

# ---------------------------------------------------------------------------
# Sheet 1 — EQUITY
# Each tuple: (report_name, path, frequency, updation_time, input, input_desc,
#             output, data_type, output_desc)
# api_number is assigned automatically (E001..).
# ---------------------------------------------------------------------------
EQUITY = [
    ("List Companies", "/stocks", "Daily", "20:00 IST",
     "search, sector, limit, offset",
     "Optional name/symbol & sector filters; pagination",
     "co_code, NSESymbol, BSECode, LegalName, ShortName, SectorName, McapClass",
     "JSON array", "Paginated equity master list"),
    ("Company Master / Detail", "/stock/{symbol}", "Daily", "20:00 IST",
     "symbol", "NSE symbol (case-insensitive)",
     "co_code, isin, NSESymbol, BSECode, CIN, LegalName, ShortName, Sector, Industry, FaceValue, ListingDate, McapClass, Website, RTAName",
     "JSON object", "Full canonical company master record"),
    ("Identifier Crosswalk", "/stock/{symbol}/identifiers", "Daily", "20:15 IST",
     "symbol", "NSE symbol",
     "isin, NSESymbol, BSECode, co_code, CIN, AMFI links, source, confidence",
     "JSON array", "All resolved identifiers for the entity"),
    ("Capital Structure", "/stock/{symbol}/capital-structure", "Weekly", "Sun 08:00 IST",
     "symbol", "NSE symbol",
     "AuthorisedCapital, PaidUpCapital, SharesOutstanding, FreeFloatShares, FreeFloatFactor",
     "JSON object", "Capital & free-float structure as-of latest"),
    ("EOD Price History", "/stock/{symbol}/prices", "Daily", "18:30 IST",
     "symbol, exchange, from, to, limit",
     "Symbol + optional exchange (NSE/BSE) & date range",
     "TradeDate, Exchange, Open, High, Low, Close, PrevClose, VWAP, TotalVolume, TotalTurnover, DeliverableQty, DeliverablePercent, NoOfTrades",
     "JSON array", "Daily OHLCV + delivery time series"),
    ("Latest Quote", "/stock/{symbol}/quote", "Daily", "18:30 IST",
     "symbol", "NSE symbol",
     "TradeDate, Close, PrevClose, ChangePct, Week52High, Week52Low, VWAP, TotalVolume",
     "JSON object", "Most recent EOD snapshot with day change"),
    ("Market EOD (Bhavcopy)", "/market/eod", "Daily", "18:30 IST",
     "date, exchange", "Trade date + exchange",
     "Per-symbol OHLCV rows for the whole market on a date",
     "JSON array", "Full bhavcopy for a single date"),
    ("Intraday Bars", "/stock/{symbol}/intraday", "Real-time", "09:15-15:30 IST",
     "symbol, interval, date", "Symbol + bar interval (1m/5m/15m/60m)",
     "Timestamp, Open, High, Low, Close, Volume, LTP, OpenInterest",
     "JSON array", "Intraday OHLC bars (roadmap)"),
    ("Market Depth", "/stock/{symbol}/depth", "Real-time", "09:15-15:30 IST",
     "symbol", "NSE symbol",
     "BidPrice/Qty x5, AskPrice/Qty x5, TotalBuyQty, TotalSellQty",
     "JSON object", "Level-5 order book snapshot (roadmap)"),
    ("Adjusted Price History", "/stock/{symbol}/prices/adjusted", "Daily", "19:30 IST",
     "symbol, from, to", "Symbol + date range",
     "TradeDate, CloseRaw, CloseSplitAdj, CloseTotalReturn, AdjustmentFactor, AdjustmentReason",
     "JSON array", "Split/bonus & dividend-adjusted price series"),
    ("52-Week & Price Stats", "/stock/{symbol}/price-stats", "Daily", "18:30 IST",
     "symbol", "NSE symbol",
     "Week52High, Week52Low, AllTimeHigh, AllTimeLow, AvgVolume30D, AvgDelivery30D",
     "JSON object", "Rolling price/volume statistics"),
    ("Price Band / Circuit", "/stock/{symbol}/price-band", "Daily", "18:00 IST",
     "symbol", "NSE symbol",
     "UpperCircuit, LowerCircuit, PriceBandPct",
     "JSON object", "Applicable circuit limits"),
    ("Corporate Actions", "/stock/{symbol}/corporate-actions", "Daily", "19:00 IST",
     "symbol, action_type",
     "Symbol + optional type (DIVIDEND/BONUS/SPLIT/...)",
     "ActionType, AnnouncementDate, ExDate, RecordDate, EffectiveDate, DividendPerShare, RatioNumerator, RatioDenominator, OldFaceValue, NewFaceValue",
     "JSON array", "All corporate actions for the entity"),
    ("Dividends History", "/stock/{symbol}/dividends", "Daily", "19:00 IST",
     "symbol", "NSE symbol",
     "ExDate, RecordDate, DividendPerShare, DividendPercent, ActionSubType",
     "JSON array", "Dividend payout history"),
    ("Splits & Bonus", "/stock/{symbol}/splits-bonus", "Daily", "19:00 IST",
     "symbol", "NSE symbol",
     "ExDate, ActionType, RatioNumerator, RatioDenominator, OldFaceValue, NewFaceValue",
     "JSON array", "Split / bonus / FV-change history"),
    ("Profit & Loss", "/stock/{symbol}/financials/pnl", "Quarterly", "Result day 21:00 IST",
     "symbol, basis, period_type",
     "Symbol + Standalone/Consolidated + Q/Annual",
     "PeriodEnd, Revenue, EBITDA, EBIT, Interest, Depreciation, PBT, Tax, PAT, EPS",
     "JSON array", "Profit & loss statement line items"),
    ("Balance Sheet", "/stock/{symbol}/financials/balance-sheet", "Quarterly", "Result day 21:00 IST",
     "symbol, basis", "Symbol + Standalone/Consolidated",
     "PeriodEnd, ShareCapital, Reserves, Borrowings, TotalLiabilities, FixedAssets, Investments, Inventories, Receivables, Cash, TotalAssets",
     "JSON array", "Balance sheet line items"),
    ("Cash Flow", "/stock/{symbol}/financials/cashflow", "Quarterly", "Result day 21:00 IST",
     "symbol, basis", "Symbol + Standalone/Consolidated",
     "PeriodEnd, CFO, CFI, CFF, NetCashFlow, OpeningCash, ClosingCash, FreeCashFlow",
     "JSON array", "Cash flow statement line items"),
    ("Quarterly Results", "/stock/{symbol}/results/quarterly", "Quarterly", "Result day 21:00 IST",
     "symbol", "NSE symbol",
     "PeriodEnd, Revenue, PAT, EPS, OPM, AuditStatus, FilingDate",
     "JSON array", "Quarterly result summary"),
    ("Financial Ratios", "/stock/{symbol}/ratios", "Quarterly", "Result day 22:00 IST",
     "symbol, basis", "Symbol + basis",
     "PeriodEnd, GrossMargin, OperatingMargin, NetMargin, ROE, ROCE, ROA, CurrentRatio, DebtToEquity, EPS, BookValuePerShare, SalesGrowthYoY",
     "JSON array", "Derived fundamental ratios"),
    ("Valuation Metrics", "/stock/{symbol}/valuation", "Daily", "19:45 IST",
     "symbol", "NSE symbol",
     "MarketCap, FreeFloatMarketCap, EnterpriseValue, PE_TTM, PB, PS, EV_EBITDA, DividendYield, PEG",
     "JSON object", "Current valuation multiples"),
    ("Shareholding Pattern", "/stock/{symbol}/shareholding", "Quarterly", "Q-end+21d 20:00 IST",
     "symbol", "NSE symbol",
     "QuarterEnd, PromoterPct, PromoterGroupPct, FIIPct, DIIPct, MutualFundPct, PublicPct, PledgedPct, NumberOfShareholders",
     "JSON array", "Quarterly shareholding pattern"),
    ("Shareholding Holders (>1%)", "/stock/{symbol}/shareholding/holders", "Quarterly", "Q-end+21d 20:00 IST",
     "symbol, quarter", "Symbol + optional quarter-end",
     "HolderType, HolderName, SharesHeld, HoldingPct, HoldingValue, ChangeQoQShares",
     "JSON array", "Named holders above 1%"),
    ("MFs Holding This Stock", "/stock/{symbol}/mf-holders", "Monthly", "Q-end+25d 20:00 IST",
     "symbol", "NSE symbol",
     "SchemeCode, SchemeName, AMCName, WeightPct, MarketValue, QuarterEnd",
     "JSON array", "Reverse mf_holding bridge: funds owning the stock"),
    ("Promoter Pledge", "/stock/{symbol}/pledge", "Quarterly", "Q-end+21d 20:00 IST",
     "symbol", "NSE symbol",
     "QuarterEnd, PromoterPledgePct, PledgedPct",
     "JSON array", "Pledge / encumbrance history"),
    ("Board & Management", "/stock/{symbol}/management", "Quarterly", "Sun 09:00 IST",
     "symbol", "NSE symbol",
     "PersonName, DIN, Designation, Category, AppointmentDate, Remuneration",
     "JSON array", "Directors & key managerial personnel"),
    ("Announcements", "/stock/{symbol}/announcements", "Real-time", "Every 15 min",
     "symbol, from, to, category", "Symbol + date range + category",
     "AnnouncementDateTime, Exchange, Category, Subject, AttachmentURL",
     "JSON array", "Exchange announcements / filings"),
    ("Events Calendar", "/stock/{symbol}/events", "Daily", "07:30 IST",
     "symbol", "NSE symbol",
     "EventType, EventDate, Purpose, ExDate, RecordDate",
     "JSON array", "Board meetings, AGM, result dates, ex-dates"),
    ("Insider Trading (PIT)", "/stock/{symbol}/insider-trades", "Daily", "19:30 IST",
     "symbol", "NSE symbol",
     "AcquirerName, Category, TransactionType, Shares, Value, TransactionDate, ModeOfAcquisition",
     "JSON array", "SEBI PIT insider transactions"),
    ("Bulk Deals", "/stock/{symbol}/bulk-deals", "Daily", "19:00 IST",
     "symbol", "NSE symbol",
     "DealDate, Exchange, ClientName, BuySell, Quantity, Price, Value",
     "JSON array", "Bulk deals on the stock"),
    ("Block Deals", "/stock/{symbol}/block-deals", "Daily", "19:00 IST",
     "symbol", "NSE symbol",
     "DealDate, Exchange, ClientName, BuySell, Quantity, Price, Value",
     "JSON array", "Block deals on the stock"),
    ("Short Selling / SLB", "/stock/{symbol}/short-selling", "Daily", "20:00 IST",
     "symbol", "NSE symbol",
     "TradeDate, ShortQty, ShortValue, SLB_OutstandingQty, LendingFeePct",
     "JSON array", "Short-sell & securities-lending data"),
    ("Surveillance Flags", "/stock/{symbol}/surveillance", "Daily", "18:00 IST",
     "symbol", "NSE symbol",
     "AsOfDate, ASMStage, GSMStage, InFnOBan, CircuitStage, SuspensionFlag",
     "JSON object", "ASM/GSM/ban/surveillance status"),
    ("Credit Ratings", "/stock/{symbol}/credit-ratings", "Weekly", "Sat 09:00 IST",
     "symbol", "NSE symbol",
     "Agency, InstrumentType, RatingScale, Rating, Outlook, RatingAction, RatingDate",
     "JSON array", "Issuer & instrument credit ratings"),
    ("Technical Indicators", "/stock/{symbol}/technicals", "Daily", "19:45 IST",
     "symbol, date", "Symbol + optional date",
     "SMA50, SMA200, EMA20, RSI14, MACD, ATR14, BollingerUpper/Lower, ADX",
     "JSON object", "Daily technical indicator set"),
    ("Pivot Levels", "/stock/{symbol}/pivots", "Daily", "19:45 IST",
     "symbol, method", "Symbol + classic/fibonacci/camarilla",
     "Pivot, R1, R2, R3, S1, S2, S3",
     "JSON object", "Support/resistance pivot levels"),
    ("Risk / Beta", "/stock/{symbol}/risk", "Daily", "19:45 IST",
     "symbol", "NSE symbol",
     "Beta1Y, Volatility30D, Volatility1Y, CorrelationNifty",
     "JSON object", "Beta & volatility measures"),
    ("Futures (F&O)", "/stock/{symbol}/derivatives/futures", "Daily", "18:30 IST",
     "symbol, expiry", "Symbol + optional expiry",
     "ExpiryDate, Open, High, Low, Close, Settle, OpenInterest, ChangeInOI, Volume",
     "JSON array", "Stock futures EOD + OI"),
    ("Option Chain", "/stock/{symbol}/derivatives/options", "Intraday", "Snapshots intraday",
     "symbol, expiry", "Symbol + expiry",
     "StrikePrice, OptionType, LTP, OpenInterest, ChangeInOI, ImpliedVolatility, Delta, Gamma, Theta, Vega",
     "JSON array", "Option chain with greeks & OI"),
    ("Peer Comparison", "/stock/{symbol}/peers", "Quarterly", "Q-end+25d IST",
     "symbol", "NSE symbol",
     "PeerCoCode, PeerName, MarketCap, PE, PB, ROE, SalesGrowth, NetMargin, RankInPeerGroup",
     "JSON array", "Side-by-side peer fundamentals"),
    ("Analyst Estimates", "/stock/{symbol}/estimates", "Weekly", "Sat 10:00 IST",
     "symbol, fiscal_year", "Symbol + FY",
     "MetricType, ConsensusMean, High, Low, NumAnalysts, TargetPrice, Recommendation, BuyCount, HoldCount, SellCount",
     "JSON array", "Consensus estimates & recommendations"),
    ("News", "/stock/{symbol}/news", "Real-time", "Every 30 min",
     "symbol, from, to", "Symbol + date range",
     "PublishedAt, Source, Headline, Summary, URL, SentimentScore",
     "JSON array", "Tagged news feed"),
    ("ESG Scores", "/stock/{symbol}/esg", "Annual", "Post-BRSR filing",
     "symbol", "NSE symbol",
     "ESGScore, EnvScore, SocialScore, GovScore, ESGRating, Provider",
     "JSON object", "ESG / sustainability scores"),
    ("List Sectors", "/sectors", "Monthly", "1st 08:00 IST",
     "-", "None",
     "SectorCode, SectorName, MacroSector, Industry, BasicIndustry, CompanyCount",
     "JSON array", "Sector taxonomy master"),
    ("Sector Constituents", "/sector/{sector_code}/companies", "Daily", "20:00 IST",
     "sector_code", "Sector code",
     "co_code, NSESymbol, LegalName, MarketCap, McapClass",
     "JSON array", "Companies in a sector"),
    ("List Indices", "/indices", "Daily", "18:30 IST",
     "index_type", "Optional broad/sector/thematic",
     "IndexCode, IndexName, IndexType, Provider",
     "JSON array", "Index master list"),
    ("Index Value History", "/index/{index_code}/values", "Daily", "18:30 IST",
     "index_code, from, to", "Index + date range",
     "TradeDate, Open, High, Low, Close, PctChange, IndexPE, IndexPB, IndexDivYield",
     "JSON array", "Index OHLC + valuation series"),
    ("Index Constituents", "/index/{index_code}/constituents", "Daily", "18:30 IST",
     "index_code", "Index code",
     "co_code, NSESymbol, Weight, FreeFloatMcap, JoinDate",
     "JSON array", "Constituents & weights"),
    ("Market Breadth", "/market/breadth", "Daily", "18:30 IST",
     "date, exchange", "Date + exchange",
     "Advances, Declines, Unchanged, New52WHigh, New52WLow, AdvanceDeclineRatio",
     "JSON object", "Market-wide breadth statistics"),
    ("FII / DII Flows", "/market/fii-dii", "Daily", "21:00 IST",
     "date", "Trade date",
     "FIINetCash, DIINetCash, FIIBuy, FIISell, DIIBuy, DIISell",
     "JSON object", "Institutional cash-market flows"),
    ("Bulk Deals (Market)", "/market/bulk-deals", "Daily", "19:00 IST",
     "date, exchange", "Date + exchange",
     "co_code, NSESymbol, ClientName, BuySell, Quantity, Price",
     "JSON array", "All bulk deals on a date"),
    ("Block Deals (Market)", "/market/block-deals", "Daily", "19:00 IST",
     "date, exchange", "Date + exchange",
     "co_code, NSESymbol, ClientName, BuySell, Quantity, Price",
     "JSON array", "All block deals on a date"),
    ("IPO Calendar", "/ipos", "Daily", "20:00 IST",
     "status", "open/upcoming/listed",
     "IssueId, co_code, IssueType, PriceBandLow, PriceBandHigh, OpenDate, CloseDate, ListingDate, IssueSize",
     "JSON array", "Primary-market issue calendar"),
    ("IPO Detail / Subscription", "/ipo/{issue_id}", "Daily", "20:00 IST",
     "issue_id", "Issue id",
     "QIBx, NIIx, Retailx, TotalSubscriptionx, GMP, ListingPrice, ListingGainPct",
     "JSON object", "IPO subscription & listing performance"),
    ("Market Movers", "/market/movers", "Daily", "18:30 IST",
     "date, type", "Date + gainers/losers/active",
     "co_code, NSESymbol, Close, ChangePct, TotalVolume",
     "JSON array", "Top gainers / losers / most active"),
    ("Trading Calendar", "/calendar", "Annual", "Dec 15 IST",
     "exchange, year", "Exchange + year",
     "CalendarDate, IsTradingDay, HolidayName, Segment",
     "JSON array", "Trading-day & holiday calendar"),
    ("Corporate Actions (Market)", "/market/corporate-actions", "Daily", "19:00 IST",
     "ex_date, action_type", "Ex-date + optional type",
     "co_code, NSESymbol, ActionType, ExDate, DividendPerShare, RatioNumerator, RatioDenominator",
     "JSON array", "All corporate actions by ex-date"),
]

# ---------------------------------------------------------------------------
# Sheet 2 — MUTUAL FUNDS
# ---------------------------------------------------------------------------
MUTUAL_FUNDS = [
    ("List Schemes", "/mutual-funds", "Daily", "22:00 IST",
     "search, amc, category, limit, offset",
     "Name/AMC/category filters + pagination",
     "SchemeCode, SchemeName, AMCName, Category, SubCategory, Plan, Option, AUM",
     "JSON array", "Paginated scheme master list"),
    ("Scheme Master / Detail", "/mutual-fund/{scheme_code}", "Daily", "22:00 IST",
     "scheme_code", "AMFI scheme code",
     "SchemeCode, ISINGrowth, ISINIDCW, SchemeName, AMCName, Category, SubCategory, Plan, Option, LaunchDate, Benchmark, ExpenseRatio, ExitLoad, FundManager, AUM",
     "JSON object", "Full scheme master record"),
    ("NAV History", "/mutual-fund/{scheme_code}/nav", "Daily", "23:00 IST",
     "scheme_code, from, to, limit", "Scheme + date range",
     "NAVDate, NAV, RepurchasePrice, SalePrice, DayChange, DayChangePct",
     "JSON array", "Daily NAV time series"),
    ("Latest NAV", "/mutual-fund/{scheme_code}/nav/latest", "Daily", "23:00 IST",
     "scheme_code", "AMFI scheme code",
     "NAVDate, NAV, DayChange, DayChangePct",
     "JSON object", "Most recent NAV"),
    ("NAV (Market) by Date", "/mf/nav", "Daily", "23:00 IST",
     "date", "NAV date",
     "SchemeCode, SchemeName, NAV",
     "JSON array", "All scheme NAVs for a date (AMFI snapshot)"),
    ("Portfolio Holdings", "/mutual-fund/{scheme_code}/holdings", "Monthly", "12th 21:00 IST",
     "scheme_code, instrument_type, limit",
     "Scheme + optional instrument type filter",
     "InstrumentName, ISIN, co_code, InstrumentType, Sector, Quantity, MarketValue, WeightPct, Rating",
     "JSON array", "Portfolio holdings (weight desc)"),
    ("Asset Allocation", "/mutual-fund/{scheme_code}/asset-allocation", "Monthly", "12th 22:00 IST",
     "scheme_code", "AMFI scheme code",
     "AsOfDate, EquityPct, DebtPct, CashPct, OtherPct, LargeCapPct, MidCapPct, SmallCapPct",
     "JSON object", "Asset-class allocation"),
    ("Sector Allocation", "/mutual-fund/{scheme_code}/sector-allocation", "Monthly", "12th 22:00 IST",
     "scheme_code", "AMFI scheme code",
     "SectorName, WeightPct, BenchmarkWeightPct, ActiveWeightPct, Rank",
     "JSON array", "Sector-wise portfolio weights"),
    ("Returns", "/mutual-fund/{scheme_code}/returns", "Daily", "23:30 IST",
     "scheme_code", "AMFI scheme code",
     "AsOfDate, Return1M, Return3M, Return6M, Return1Y, Return3Y, Return5Y, ReturnSI, BenchmarkReturn, CategoryReturn",
     "JSON object", "Point-to-point & trailing returns"),
    ("Rolling Returns", "/mutual-fund/{scheme_code}/returns/rolling", "Weekly", "Sat 08:00 IST",
     "scheme_code, window", "Scheme + window (1Y/3Y/5Y)",
     "Date, RollingReturn, BenchmarkRollingReturn",
     "JSON array", "Rolling-return series"),
    ("SIP Returns (XIRR)", "/mutual-fund/{scheme_code}/returns/sip", "Daily", "23:30 IST",
     "scheme_code", "AMFI scheme code",
     "SIPReturn1Y, SIPReturn3Y, SIPReturn5Y",
     "JSON object", "SIP XIRR over standard tenors"),
    ("Risk Metrics", "/mutual-fund/{scheme_code}/risk", "Weekly", "Sat 08:00 IST",
     "scheme_code, period", "Scheme + period (3Y/5Y)",
     "StdDev, Sharpe, Sortino, Beta, Alpha, RSquared, Treynor, MaxDrawdown",
     "JSON object", "Risk-adjusted performance metrics"),
    ("AUM History", "/mutual-fund/{scheme_code}/aum", "Monthly", "8th 10:00 IST",
     "scheme_code", "AMFI scheme code",
     "MonthEnd, AUM, AAUM, FolioCount",
     "JSON array", "Scheme AUM time series"),
    ("Expense & Load", "/mutual-fund/{scheme_code}/expense", "Monthly", "5th 10:00 IST",
     "scheme_code", "AMFI scheme code",
     "AsOfDate, TERRegular, TERDirect, ExitLoadSlab, EntryLoad",
     "JSON object", "Expense ratio & load structure"),
    ("IDCW History", "/mutual-fund/{scheme_code}/idcw", "Event-driven", "Record day 22:00 IST",
     "scheme_code", "AMFI scheme code",
     "RecordDate, IDCWPerUnit, IDCWFrequency",
     "JSON array", "Dividend / IDCW payout history"),
    ("Fund Manager", "/mutual-fund/{scheme_code}/manager", "Monthly", "1st 09:30 IST",
     "scheme_code", "AMFI scheme code",
     "ManagerName, Qualification, ExperienceYears, ManagingSince, SchemesManagedCount",
     "JSON array", "Fund manager(s) on the scheme"),
    ("Portfolio Analytics", "/mutual-fund/{scheme_code}/portfolio-analytics", "Monthly", "13th 09:00 IST",
     "scheme_code", "AMFI scheme code",
     "PortfolioTurnoverPct, NumberOfHoldings, Top10ConcentrationPct, ActiveSharePct",
     "JSON object", "Concentration & turnover analytics"),
    ("Holdings Overlap", "/mutual-fund/overlap", "Monthly", "13th 09:00 IST",
     "a, b", "Two scheme codes",
     "OverlapPct, CommonHoldings[], SchemeAExclusive[], SchemeBExclusive[]",
     "JSON object", "Portfolio overlap between two schemes"),
    ("Debt Analytics", "/mutual-fund/{scheme_code}/debt-analytics", "Monthly", "13th 09:00 IST",
     "scheme_code", "AMFI scheme code",
     "AvgMaturityYears, ModifiedDuration, YTM, CreditAAAPct, SovereignPct, GSecPct, CorpBondPct",
     "JSON object", "Debt portfolio analytics"),
    ("Categorisation", "/mutual-fund/{scheme_code}/category", "Monthly", "1st 09:00 IST",
     "scheme_code", "AMFI scheme code",
     "SEBICategory, SEBISubCategory, RiskometerLevel, IsELSS, IsIndexFund, IsETF",
     "JSON object", "SEBI categorisation & flags"),
    ("Ratings & Rankings", "/mutual-fund/{scheme_code}/ratings", "Monthly", "5th 11:00 IST",
     "scheme_code", "AMFI scheme code",
     "CRISILRank, ValueResearchStars, MorningstarRating, QuartileRank",
     "JSON object", "Third-party ratings & quartile rank"),
    ("List AMCs", "/amcs", "Monthly", "1st 09:00 IST",
     "-", "None",
     "AMCCode, AMCName, SponsorName, TotalAUM, AUMRank",
     "JSON array", "AMC master list"),
    ("AMC Detail", "/amc/{amc_code}", "Monthly", "1st 09:00 IST",
     "amc_code", "AMC code",
     "AMCName, SponsorName, TrusteeName, SEBIRegNo, RTAName, CustodianName, TotalAUM",
     "JSON object", "AMC master record"),
    ("AMC Schemes", "/amc/{amc_code}/schemes", "Daily", "22:00 IST",
     "amc_code", "AMC code",
     "SchemeCode, SchemeName, Category, SubCategory, AUM",
     "JSON array", "Schemes managed by an AMC"),
    ("List Fund Managers", "/fund-managers", "Monthly", "1st 09:30 IST",
     "search", "Optional name filter",
     "ManagerId, ManagerName, ExperienceYears, SchemesManagedCount",
     "JSON array", "Fund manager master"),
    ("Fund Manager Detail", "/fund-manager/{manager_id}", "Monthly", "1st 09:30 IST",
     "manager_id", "Manager id",
     "ManagerName, Qualification, ExperienceYears, SchemesManaged[], AvgTenureReturn",
     "JSON object", "Manager profile & track record"),
    ("NFO List", "/nfos", "Daily", "20:30 IST",
     "status", "open/upcoming/closed",
     "SchemeCode, SchemeName, AMCName, NFOOpenDate, NFOCloseDate, NFOPrice",
     "JSON array", "New fund offers"),
    ("List Categories", "/mf/categories", "Monthly", "1st 09:00 IST",
     "-", "None",
     "Category, SubCategory, SchemeCount, AvgReturn1Y",
     "JSON array", "SEBI category list with aggregates"),
    ("Category Leaderboard", "/mf/category/{category}/schemes", "Daily", "23:30 IST",
     "category, sort", "Category + sort metric",
     "SchemeCode, SchemeName, Return1Y, Return3Y, AUM, QuartileRank",
     "JSON array", "Ranked schemes within a category"),
    ("Taxation", "/mutual-fund/{scheme_code}/taxation", "Annual", "Post-Budget",
     "scheme_code", "AMFI scheme code",
     "TaxBucket, LTCGHoldingMonths, STCGRate, LTCGRate, IndexationEligible",
     "JSON object", "Applicable taxation rules"),
    ("Compare Schemes", "/mf/compare", "Daily", "23:30 IST",
     "codes", "Comma-separated scheme codes",
     "Per-scheme: Return1Y/3Y/5Y, ExpenseRatio, AUM, Sharpe, StdDev, QuartileRank",
     "JSON array", "Side-by-side scheme comparison"),
    ("Stock Exposure (Reverse)", "/mf/stock-exposure/{symbol}", "Monthly", "Q-end+25d 20:00 IST",
     "symbol", "NSE symbol",
     "SchemeCode, SchemeName, AMCName, WeightPct, MarketValue",
     "JSON array", "All MF schemes holding a given stock"),
]

# ---------------------------------------------------------------------------
# Sheet 3 — CMOTS COMMENTS
# Columns: Remark, Cmots Comments
# ---------------------------------------------------------------------------
CMOTS_COMMENTS = [
    ("Canonical entity key is co_code (ISIN-backed), mirroring CMOTS co_code.", ""),
    ("Company Master V1 sector coverage frozen at 75.6%; not optimised further.", ""),
    ("EOD prices sourced NSE+BSE bhavcopy; per-exchange rows + consolidated view.", ""),
    ("Corporate actions dedup across NSE CA API + BSE announcements; ratios parsed.", ""),
    ("Shareholding from NSE master + XBRL; BSE has no clean public endpoint.", ""),
    ("MF holdings parser is generic/AMC-agnostic; proven on SBI + ICICI.", ""),
    ("mf_holding.co_code is the equity<->MF bridge powering reverse-holding queries.", ""),
    ("LIVE today: Company Master, Equity EOD, Corp Actions, Shareholding, MF Scheme, MF NAV, MF Holdings.", ""),
    ("Financial statements & ratios are NEXT; sourced via XBRL with Screener fallback.", ""),
    ("Intraday/real-time & F&O are ROADMAP; require broker WebSocket licensing.", ""),
    ("Refresh windows tuned to NSE/BSE bhavcopy (~18:30) and AMFI NAV (~23:00) IST.", ""),
    ("All gold output also materialised as Parquet for the API layer (lazy scan).", ""),
    ("Adjusted prices derived in-house from EOD + corporate actions, not sourced.", ""),
    ("Identifier crosswalk keyed on ISIN; resolver confidence + reason are tracked.", ""),
    ("Storage is SQLite for V1; same DDL ports to Postgres for scale-out.", ""),
    ("Lineage: every bronze write logged in ingest_manifest (run_id, sha256, rows).", ""),
    ("API URLs are versioned (/api/v1); routers currently mount at root in FastAPI app.", ""),
    ("Analyst estimates / ESG / credit ratings depend on licensed or scraped sources.", ""),
    ("Risk metrics & returns for MF computed in-house from NAV + benchmark + risk-free.", ""),
    ("Open items for CMOTS review: F&O depth, news licensing, ESG provider choice.", ""),
]


def _style_header(ws, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _write_api_sheet(ws, rows, prefix: str) -> None:
    headers = [h for _, h, _ in API_COLUMNS]
    ws.append(headers)
    body_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, row in enumerate(rows, start=1):
        report, path, freq, upd, inp, inp_d, out, dtype, out_d = row
        api_no = f"{prefix}{i:03d}"
        ws.append([api_no, report, f"{BASE}{path}", freq, upd, inp, inp_d, out, dtype, out_d])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = body_align
            cell.border = border
    for idx, (_, _, width) in enumerate(API_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _style_header(ws, len(headers))


def _write_comments_sheet(ws, rows) -> None:
    ws.append(["Remark", "Cmots Comments"])
    body_align = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for remark, comment in rows:
        ws.append([remark, comment])
    for r in range(2, ws.max_row + 1):
        for c in (1, 2):
            cell = ws.cell(row=r, column=c)
            cell.alignment = body_align
            cell.border = border
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 60
    _style_header(ws, 2)


def build_workbook(out_path: Path) -> Path:
    wb = Workbook()
    ws_eq = wb.active
    ws_eq.title = "EQUITY"
    _write_api_sheet(ws_eq, EQUITY, "E")

    ws_mf = wb.create_sheet("MUTUAL FUNDS")
    _write_api_sheet(ws_mf, MUTUAL_FUNDS, "M")

    ws_cm = wb.create_sheet("CMOTS COMMENTS")
    _write_comments_sheet(ws_cm, CMOTS_COMMENTS)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    out = Path(__file__).with_name("CMOTS_Replacement_API_Catalog.xlsx")
    path = build_workbook(out)
    print(f"Wrote {path}")
    print(f"  EQUITY        : {len(EQUITY)} APIs (E001..E{len(EQUITY):03d})")
    print(f"  MUTUAL FUNDS  : {len(MUTUAL_FUNDS)} APIs (M001..M{len(MUTUAL_FUNDS):03d})")
    print(f"  CMOTS COMMENTS: {len(CMOTS_COMMENTS)} remarks")
