import polars as pl
import openpyxl

wb = openpyxl.load_workbook("scratch_sbi_infra_portfolio.xlsx", read_only=True)
print("Sheet Names:", wb.sheetnames)

for sheet in wb.sheetnames[:5]:
    print(f"\n--- Sheet: {sheet} ---")
    # Read sheet with openpyxl and print first 15 rows
    ws = wb[sheet]
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx >= 20:
            break
        print(f"Row {idx}: {row}")
