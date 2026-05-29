"""MF Holdings normalizer — parses SBI monthly portfolio Excel files.

Parses Excel sheets, standardizes headers, cleans data, splits Sector and CreditRating,
classifies InstrumentType, and replicates holdings to AMFI SchemeCodes.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date
import json
import re
from pathlib import Path
import openpyxl
import polars as pl

from ..core.errors import ParseError
from ..core.logging import get_logger
from ..core.settings import load_settings

log = get_logger("normalize.mf_holdings")

HEADER_ALIASES = {
    "holding_isin": ["isin", "isincode", "isinno", "isinnumber"],
    "holding_name": [
        "nameoftheinstrumentissuer", "nameoftheinstrument", "issuersecurityname",
        "issuernameoftheinstrument", "securityname", "instrumentname",
        "companyissuerinstrumentname",  # ICICI
    ],
    "quantity": ["quantity", "qty", "numberofshares", "sharesquantity"],
    "market_value": [
        "marketvaluerisinlakhs", "marketfairvaluerisinlakhs", "marketvaluerslakhs",
        "marketvaluerinlakhs", "marketvalue", "fairvalue",
        "exposuremarketvaluerslakh",  # ICICI
    ],
    "weight_pct": ["toaum", "tonetassets", "tonav", "weight", "weightpct", "allocationpercentage"],
    "sector": ["ratingindustry", "industryrating", "sectorrating", "industry", "rating", "sector"],
}


def normalize_header(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def clean_for_mapping(name: str | None) -> str:
    if not name:
        return ""
    name = name.lower()
    name = name.replace("and", "").replace("&", "")
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"\b(formerly known as|formerly)\b.*$", "", name)
    name = re.sub(r"\b(direct|regular|plan|growth|dividend|idcw|option|payout|reinvestment|reinvest)\b", "", name)
    name = re.sub(r"[^a-z0-9]", "", name)
    return name


_PLAN_TOKENS = {"direct", "regular"}
_OPTION_TOKENS = {"growth", "idcw", "dividend", "reinvestment", "payout", "bonus"}


def _signature_tokens(name: str | None) -> tuple[frozenset, frozenset]:
    """(plan tokens, option tokens) present in a fund/scheme name."""
    words = re.findall(r"[a-z]+", (name or "").lower())
    return (frozenset(w for w in words if w in _PLAN_TOKENS),
            frozenset(w for w in words if w in _OPTION_TOKENS))


def _signature_match(fund_name: str, scheme_name: str) -> bool:
    """Disambiguate a scheme contested by multiple funds sharing a normalized key.

    A fund file with no plan/option token is fund-level (matches any plan); otherwise
    its plan/option must intersect the scheme's. Generic, no AMC-specific logic.
    """
    fp, fo = _signature_tokens(fund_name)
    sp, so = _signature_tokens(scheme_name)
    plan_ok = (not fp) or (not sp) or bool(fp & sp)
    opt_ok = (not fo) or (not so) or bool(fo & so)
    return plan_ok and opt_ok


def classify_instrument_type(
    holding_name: str, sector: str | None, current_section: str
) -> str:
    sec_title = current_section.upper()

    # P0 #1: the EQUITY section is authoritative. Return immediately and NEVER
    # apply name-substring heuristics here — a 2-letter substring like "cp"/"cd"
    # (e.g. CDSL) or "gold"/"silver"/"trust" (e.g. Goldiam) would otherwise
    # misclassify a real equity as DEBT/CASH/commodity and drop its co_code.
    if sec_title == "EQUITY":
        return "EQUITY"

    # Substring-based classification runs ONLY for non-equity sections.
    name = holding_name.lower()

    if "unlisted" in name:
        return "UNKNOWN"
    if "commercial paper" in name or "cp" in name or sec_title == "COMMERCIAL_PAPER":
        return "DEBT"
    if "certificate of deposit" in name or "cd" in name or sec_title == "CERTIFICATE_OF_DEPOSITS":
        return "DEBT"
    if "treps" in name or "cblo" in name or "reverse repo" in name or sec_title == "TREPS":
        return "TREPS"
    if (
        "cash & cash equivalents" in name
        or "net current assets" in name
        or "net receivable" in name
        or "margin amount" in name
        or sec_title == "CASH"
    ):
        return "CASH"

    if sec_title in ("DEBT", "MONEY_MARKET", "COMMERCIAL_PAPER",
                     "CERTIFICATE_OF_DEPOSITS", "TREASURY_BILLS"):
        return "DEBT"
    if "reit" in name or "invit" in name or "trust" in name or (
        sec_title == "OTHERS" and ("reit" in name or "invit" in name)
    ):
        if "reit" in name:
            return "REIT"
        else:
            return "INVIT"
    if "etf" in name or "bees" in name:
        return "ETF"
    if "gold" in name:
        return "GOLD"
    if "silver" in name:
        return "SILVER"

    return "UNKNOWN"


# --- Generic structural detection (breakpoints #2 & #3) --------------------------
# A *holding* carries a security identifier (an ISIN and/or a quantity). Rows that
# carry only an aggregate name/value are section banners, sub-banners or subtotals.
# These keyword lists describe the vocabulary of Indian MF factsheets in general —
# NOT any single AMC's layout — so the same logic serves SBI, ICICI and beyond.

# name-substring -> section. SKIP entries first so e.g. "Equity Derivatives" skips.
_SECTION_RULES = [
    ("derivative", "SKIP"), ("hedging position", "SKIP"), ("futures", "SKIP"),
    ("options", "SKIP"), ("benchmark", "SKIP"), ("net asset value", "SKIP"),
    ("floating rate", "SKIP"), ("deviation", "SKIP"),
    ("equity & equity", "EQUITY"), ("equity and equity", "EQUITY"),
    ("equity shares", "EQUITY"), ("equity related", "EQUITY"),
    ("listed / awaiting", "EQUITY"), ("listed/awaiting", "EQUITY"),
    ("awaiting listing", "EQUITY"), ("unlisted", "EQUITY"),
    ("commercial paper", "COMMERCIAL_PAPER"),
    ("certificate of deposit", "CERTIFICATE_OF_DEPOSITS"),
    ("treasury bill", "TREASURY_BILLS"),
    ("government securit", "DEBT"), ("g-sec", "DEBT"), ("sovereign", "DEBT"),
    ("money market", "MONEY_MARKET"),
    ("debt instrument", "DEBT"), ("debt securit", "DEBT"),
    ("non convertible", "DEBT"), ("non-convertible", "DEBT"),
    ("corporate debt", "DEBT"), ("bonds and ncd", "DEBT"), ("bonds & ncd", "DEBT"),
    # Cash bucket (current-assets / receivables-payables vocabulary).
    ("other current asset", "CASH"), ("other current liabilit", "CASH"),
    ("current asset", "CASH"), ("current liabilit", "CASH"),
    ("net receivable", "CASH"), ("net payable", "CASH"),
    ("receivable", "CASH"), ("payable", "CASH"),
    ("mutual fund unit", "ETF"), ("units of mutual fund", "ETF"),
    ("exchange traded fund", "ETF"),
]

# Identifier-less rows worth keeping as holdings (cash-bucket instruments).
_KEEP_INSTRUMENT_KEYWORDS = (
    "treps", "reverse repo", "triparty repo", "tri-party repo",
    "net current asset", "net receivable", "net payable", "margin",
    "cash and cash equivalent", "cash & cash equivalent",
    "clearing corporation", "corporate debt market development",
)

# Identifier-less rows to drop (subtotals / notes / metadata / NAV / footers).
_NOISE_KEYWORDS = (
    "total", "sub total", "subtotal", "notes", "note -", "note-",
    "% to", "% of", "portfolio as on", "portfolio statement",
    "plan name", "scheme name", "face value", "record date",
    "yield of the", "disclaimer", "nav per unit", "average maturity",
    "macaulay", "modified duration", "ytm", "annualised", "expense ratio",
)


def _detect_section(name: str) -> str | None:
    n = name.lower()
    for kw, sec in _SECTION_RULES:
        if kw in n:
            return sec
    return None


def _is_keep_instrument(name: str) -> bool:
    n = name.lower()
    return any(k in n for k in _KEEP_INSTRUMENT_KEYWORDS)


def _is_noise(name: str) -> bool:
    n = name.lower()
    return any(k in n for k in _NOISE_KEYWORDS)


# A real security identifier is a 12-char ISIN (2 letters + 9 alnum + check digit).
# Rows carrying a stray number in the ISIN column (NAV/plan footers) therefore have
# "no identifier" and fall into the structural-row gate — generic, not AMC-specific.
_ISIN_RE = re.compile(r"[A-Z]{2}[A-Z0-9]{9}[0-9]")


def _looks_like_isin(v) -> bool:
    if v is None:
        return False
    return bool(_ISIN_RE.fullmatch(str(v).strip().upper()))


def load_amfi_schemes(as_of: Path | pl.DataFrame | str, amc_filter: str = "") -> pl.DataFrame:
    # If a path was passed, find the nearest AMFI txt file to parse
    settings = load_settings()
    partition_date = None
    # Extract date from path
    if isinstance(as_of, Path):
        m = re.search(r"dt=(\d{4}-\d{2}-\d{2})", str(as_of))
        if m:
            partition_date = m.group(1)
    if not partition_date:
        partition_date = date.today().isoformat()

    raw_dir = settings.resolve(Path(f"storage/raw/amfi/navall/dt={partition_date}"))
    if not raw_dir.exists():
        log.warning("normalize.mf_holdings.amfi_raw_not_found", dir=str(raw_dir))
        return pl.DataFrame([])
    txt_files = list(raw_dir.glob("*.txt"))
    if not txt_files:
        log.warning("normalize.mf_holdings.amfi_txt_not_found", dir=str(raw_dir))
        return pl.DataFrame([])
    try:
        from ..normalizers.mf_nav import parse_navall_file
        df = parse_navall_file(txt_files[0])
        # Filter to the AMC under processing (token supplied by the manifest/adapter).
        if amc_filter:
            df = df.filter(pl.col("amc_name").str.contains(amc_filter))
        return df
    except Exception as e:
        log.warning("normalize.mf_holdings.amfi_parse_failed", error=str(e))
        return pl.DataFrame([])


def parse_portfolio_file(path: Path) -> pl.DataFrame:
    """Read raw SBI portfolio directory containing Excel files and return standardized silver DataFrame."""
    log.info("normalize.mf_holdings.start", path=str(path))

    silver_schema = {
        "scheme_code": pl.Int64,
        "scheme_name": pl.Utf8,
        "amc_name": pl.Utf8,
        "holding_name": pl.Utf8,
        "holding_isin": pl.Utf8,
        "weight_pct": pl.Float64,
        "quantity": pl.Float64,
        "market_value": pl.Float64,
        "sector": pl.Utf8,
        "credit_rating": pl.Utf8,
        "instrument_type": pl.Utf8,
        "quarter_end": pl.Date,
        "source": pl.Utf8,
    }

    manifest_file = path / "manifest.json"
    if not manifest_file.exists():
        raise ParseError(f"MFHoldings: manifest.json missing at {manifest_file}")

    try:
        manifest = json.loads(manifest_file.read_text(encoding="utf-8"))
    except Exception as e:
        raise ParseError(f"Failed to read manifest.json: {e}")

    # AMC identity is supplied by the adapter via the manifest — no hardcoding.
    source = manifest.get("source") or "UNKNOWN"
    amc_name = manifest.get("amc") or "Unknown AMC"
    amfi_filter = manifest.get("amfi_filter") or source

    # Load AMFI schemes for SchemeCode mapping.
    # P0 #2: an empty AMFI map means NO holding can resolve to a SchemeCode —
    # the whole run would silently produce zero rows. Fail loudly instead.
    amfi_df = load_amfi_schemes(path, amfi_filter)
    if amfi_df.height == 0:
        raise ParseError(
            "MFHoldings: AMFI scheme mapping is empty — cannot resolve scheme codes "
            "(check that AMFI NAVAll was ingested for this partition)"
        )

    parsed_rows = []

    for file_info in manifest.get("files", []):
        if file_info.get("download_status") != "success":
            continue

        file_name = file_info["file_name"]
        file_path = path / file_name
        # P0 #2: a success-status manifest file that is missing or unreadable is a
        # silent failure — fail loudly rather than skip.
        if not file_path.exists():
            raise ParseError(f"MFHoldings: manifest file marked success but missing on disk: {file_path}")

        log.info("normalize.mf_holdings.parsing_file", file=file_name)
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
            ws = wb.active
        except Exception as e:
            raise ParseError(f"MFHoldings: failed to load workbook {file_name}: {e}")

        rows_before_file = len(parsed_rows)

        fund_name = None
        reporting_date = None

        # Pass 1: Extract Metadata (Scheme Name and Statement Date)
        for row in ws.iter_rows(values_only=True):
            row_str = [str(c).strip() for c in row if c is not None]
            if not row_str:
                continue
            # Look for Scheme Name
            for idx, token in enumerate(row_str):
                if "scheme name" in token.lower() and idx + 1 < len(row_str):
                    fund_name = row_str[idx + 1]
                if "portfolio statement as on" in token.lower() and idx + 1 < len(row_str):
                    reporting_date_str = row_str[idx + 1]
                    # Parse date if possible
                    # E.g. '2026-04-30 00:00:00'
                    try:
                        m_date = re.search(r"(\d{4}-\d{2}-\d{2})", reporting_date_str)
                        if m_date:
                            reporting_date = date.fromisoformat(m_date.group(1))
                    except Exception:
                        pass

        if not fund_name:
            # Fall back to file info
            fund_name = file_info["fund_name"]
        if not reporting_date:
            # Fall back to partition date from manifest
            reporting_date = date.fromisoformat(manifest["partition"])

        log.info("normalize.mf_holdings.metadata", fund=fund_name, date=str(reporting_date))

        # Pass 2: Discover headers & data rows
        col_mappings = {}
        header_row_found = False

        current_section = "UNKNOWN"

        for row_cells in ws.iter_rows(values_only=True):
            non_none = [c for c in row_cells if c is not None]
            if not non_none:
                continue

            # Header detection
            if not header_row_found:
                temp_mappings = {}
                for idx, cell in enumerate(row_cells):
                    norm_cell = normalize_header(cell)
                    for canonical_key, aliases in HEADER_ALIASES.items():
                        if norm_cell in aliases:
                            temp_mappings[canonical_key] = idx

                # Match criteria: at least 3 critical keys
                if (
                    len(temp_mappings) >= 3
                    and "holding_name" in temp_mappings
                    and "weight_pct" in temp_mappings
                    and ("holding_isin" in temp_mappings or "quantity" in temp_mappings)
                ):
                    col_mappings = temp_mappings
                    header_row_found = True
                    log.info("normalize.mf_holdings.header_discovered", cols=col_mappings)
                continue

            # ---- Generic structural detection (breakpoints #2 & #3) ----
            # Read identifiers first, then decide section vs banner vs holding.
            name_idx = col_mappings.get("holding_name")
            isin_idx = col_mappings.get("holding_isin")
            sector_idx = col_mappings.get("sector")
            qty_idx = col_mappings.get("quantity")
            mv_idx = col_mappings.get("market_value")
            wt_idx = col_mappings.get("weight_pct")

            raw_name = row_cells[name_idx] if name_idx is not None else None
            if not raw_name:
                continue
            raw_name_str = str(raw_name).strip()
            if not raw_name_str:
                continue

            raw_isin = row_cells[isin_idx] if isin_idx is not None else None
            raw_sector = row_cells[sector_idx] if sector_idx is not None else None
            raw_qty = row_cells[qty_idx] if qty_idx is not None else None
            raw_mv = row_cells[mv_idx] if mv_idx is not None else None
            raw_wt = row_cells[wt_idx] if wt_idx is not None else None

            def is_nil(v):
                if v is None:
                    return True
                s = str(v).strip().upper()
                return s in ("", "NIL", "N.A.", "NA", "-")

            has_isin = _looks_like_isin(raw_isin)
            has_qty = not is_nil(raw_qty)

            # A priced holding carries a security identifier (ISIN and/or quantity).
            # A row with neither is structural — works for SBI's single-cell banners
            # AND ICICI's subtotal-style banners that carry only a value/weight:
            #   * matches a section phrase            -> set current_section, skip the banner
            #   * under a skipped section             -> skip
            #   * subtotal / note / metadata / footer -> skip
            #   * cash-bucket instrument (TREPS/NCA)  -> keep (classified by name)
            if not has_isin and not has_qty:
                has_value = (not is_nil(raw_mv)) or (not is_nil(raw_wt))
                sec = _detect_section(raw_name_str)
                if sec is not None:
                    current_section = sec
                    # A cash-section line is usually the holding itself (Net Receivable
                    # / Payable, current assets) carrying only a value — keep those.
                    # Every other section phrase is a pure header whose priced holdings
                    # follow with their own identifiers, and value-less rows are dividers.
                    if not (sec == "CASH" and has_value):
                        continue
                elif current_section == "SKIP":
                    continue
                # Keep only a value-bearing cash-bucket instrument (TREPS/NCA); a
                # value-less or noise name is a divider/footer and is dropped.
                elif _is_noise(raw_name_str) or not has_value or not _is_keep_instrument(raw_name_str):
                    continue
            elif current_section == "SKIP":
                # priced rows beneath a skipped section (e.g. derivatives)
                continue

            # Clean and parse numbers
            def parse_float(v):
                if is_nil(v):
                    return None
                s = str(v).strip().replace(",", "")
                if s == "#":
                    return 0.0  # Less than 0.005%, represent as 0.0
                try:
                    return float(s)
                except ValueError:
                    return None

            qty = parse_float(raw_qty)
            mv = parse_float(raw_mv)
            wt = parse_float(raw_wt)

            itype = classify_instrument_type(raw_name_str, raw_sector, current_section)

            sector_val = None
            rating_val = None
            if itype == "EQUITY":
                sector_val = str(raw_sector).strip() if raw_sector else None
            elif itype == "DEBT":
                rating_val = str(raw_sector).strip() if raw_sector else None

            isin_clean = str(raw_isin).strip().upper() if _looks_like_isin(raw_isin) else None

            parsed_rows.append({
                "fund_name": fund_name,
                "holding_name": raw_name_str,
                "holding_isin": isin_clean,
                "weight_pct": wt or 0.0,
                "quantity": qty,
                "market_value": mv,
                "sector": sector_val,
                "credit_rating": rating_val,
                "instrument_type": itype,
                "quarter_end": reporting_date,
                "source": source,
            })

        # P0 #2: every successfully-loaded file must yield a header and >= 1 row.
        if not header_row_found:
            raise ParseError(f"MFHoldings: header row not found in {file_name}")
        if len(parsed_rows) == rows_before_file:
            raise ParseError(f"MFHoldings: {file_name} produced 0 parsed rows")

        # Breakpoint #4: scheme-level weight-unit auto-detection. One workbook == one
        # scheme; if its weights sum to ~1 they are fractions (ICICI) -> scale to
        # percent. If they already sum to ~100 (SBI) leave unchanged. No AMC rules.
        file_rows = parsed_rows[rows_before_file:]
        wt_sum = sum((r["weight_pct"] or 0.0) for r in file_rows)
        if 0.0 < wt_sum < 5.0:
            for r in file_rows:
                r["weight_pct"] = (r["weight_pct"] or 0.0) * 100.0

    if not parsed_rows:
        raise ParseError("MFHoldings: no holdings parsed from any file")

    # ---- Matching layer: assign each AMFI scheme to AT MOST ONE fund. ----
    # clean_for_mapping can collapse distinct funds (e.g. "Regular Savings Fund" vs
    # "Savings Fund") or two per-plan workbooks of one fund onto a shared key. Blindly
    # fanning a portfolio onto every match double-counts a scheme (~200% weight). So:
    # a uniquely-claimed scheme goes to its fund; a contested scheme is awarded only to
    # the fund whose plan/option signature uniquely matches; otherwise it is left
    # unresolved and warned. A portfolio is never duplicated onto another fund's scheme.
    funds = list({r["fund_name"] for r in parsed_rows})
    key_to_funds: dict[str, list[str]] = defaultdict(list)
    for f in funds:
        key_to_funds[clean_for_mapping(f)].append(f)

    fund_to_schemes: dict[str, list[tuple]] = defaultdict(list)
    for item in amfi_df.to_dicts():
        sc, sname = item["scheme_code"], item["scheme_name"]
        claimants = key_to_funds.get(clean_for_mapping(sname), [])
        if not claimants:
            continue
        if len(claimants) == 1:
            owner = claimants[0]
        else:
            winners = [f for f in claimants if _signature_match(f, sname)]
            if len(winners) == 1:
                owner = winners[0]
            else:
                log.warning("normalize.mf_holdings.ambiguous_scheme",
                            scheme=sname, candidates=claimants)
                continue
        fund_to_schemes[owner].append((sc, sname))

    for f in funds:
        if not fund_to_schemes.get(f):
            log.warning("normalize.mf_holdings.unmapped_scheme", fund=f)

    final_records = []
    for row in parsed_rows:
        for scheme_code, scheme_name in fund_to_schemes.get(row["fund_name"], []):
            final_records.append({
                "scheme_code": scheme_code,
                "scheme_name": scheme_name,
                "amc_name": amc_name,
                "holding_name": row["holding_name"],
                "holding_isin": row["holding_isin"],
                "weight_pct": row["weight_pct"],
                "quantity": row["quantity"],
                "market_value": row["market_value"],
                "sector": row["sector"],
                "credit_rating": row["credit_rating"],
                "instrument_type": row["instrument_type"],
                "quarter_end": row["quarter_end"],
                "source": row["source"],
            })

    # P0 #2: parsed holdings but matched zero AMFI schemes => mapping is broken.
    if not final_records:
        raise ParseError(
            "MFHoldings: parsed holdings but 0 scheme matches after AMFI mapping "
            "(fund-name normalization may not align with AMFI scheme names)"
        )

    df_out = pl.DataFrame(final_records, schema=silver_schema)
    return df_out
