"""Inspect CompanyMaster output: shape, sample rows, short-name spot checks."""
from pathlib import Path
import openpyxl

xlsx = Path("storage/output/company_master_2026-05-27.xlsx")
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]
print("sheet:", ws.title, "rows:", ws.max_row, "cols:", ws.max_column)
print("HEADER:", hdr)

idx = {h: i for i, h in enumerate(hdr)}

# Rows that are dual-listed (have both BSE + NSE) — most informative
print("\n-- 8 dual-listed samples --")
shown = 0
for r in rows[1:]:
    if r[idx["NSESymbol"]] and r[idx["BSECode"]] and shown < 8:
        print(r)
        shown += 1

# Short-name spot checks
print("\n-- short-name checks --")
targets = ["Mahindra", "Glaxo", "Deepak Fert", "Tata Consultancy", "Larsen", "ITC"]
for r in rows[1:]:
    name = (r[idx["CompanyName"]] or "")
    if any(t.lower() in name.lower() for t in targets):
        print(f"{name!r:55} -> {r[idx['CompanyShortName']]!r}")

# Sector coverage
mapped = sum(1 for r in rows[1:] if r[idx["SectorCode"]] is not None)
print(f"\nsector mapped: {mapped}/{len(rows)-1}")

# Distinct sample of mapped sectors
seen = {}
for r in rows[1:]:
    sc, sn = r[idx["SectorCode"]], r[idx["SectorName"]]
    if sc is not None and sc not in seen:
        seen[sc] = sn
print("distinct sectors present:", len(seen))
for sc in sorted(seen)[:12]:
    print(f"  {sc}: {seen[sc]}")
