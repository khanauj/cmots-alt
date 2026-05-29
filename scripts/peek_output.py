"""One-shot verifier: open the produced xlsx and print head/tail."""
from pathlib import Path
import openpyxl
import sys

xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("storage/output/mf_nav_2026-05-27.xlsx")
print(f"file: {xlsx}")
print(f"size: {xlsx.stat().st_size:,} bytes")
wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
ws = wb.active
print(f"sheet: {ws.title}  rows={ws.max_row}  cols={ws.max_column}")
rows = list(ws.iter_rows(values_only=True))
print("HEADER:")
print(" ", rows[0])
print("FIRST 3 DATA ROWS:")
for r in rows[1:4]:
    print(" ", r)
print("LAST 2 DATA ROWS:")
for r in rows[-2:]:
    print(" ", r)
